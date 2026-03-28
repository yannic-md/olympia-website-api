from openpyxl import Workbook

# Create countries Excel
wb = Workbook()
ws = wb.active
ws['A1'] = 'code'
ws['B1'] = 'name'
data = [
    ('us', 'United States'),
    ('de', 'Germany'),
    ('fr', 'France'),
    ('gb', 'Great Britain'),
    ('jp', 'Japan'),
    ('cn', 'China'),
    ('au', 'Australia'),
    ('ca', 'Canada'),
    ('it', 'Italy'),
    ('es', 'Spain'),
]
for idx, (code, name) in enumerate(data, 2):
    ws[f'A{idx}'] = code
    ws[f'B{idx}'] = name
wb.save('countries_test.xlsx')

# Create athletes Excel
wb2 = Workbook()
ws2 = wb2.active
ws2['A1'] = 'firstName'
ws2['B1'] = 'lastName'
ws2['C1'] = 'countryCode'
athletes = [
    ('Mikaela', 'Shiffrin', 'us'),
    ('Marco', 'Odermatt', 'ch'),
    ('Petra', 'Vlhova', 'sk'),
    ('Alexis', 'Pinturault', 'fr'),
    ('Sofia', 'Goggia', 'it'),
    ('Johannes', 'Boe', 'no'),
    ('Marte', 'Roeiseland', 'no'),
    ('Nathan', 'Chen', 'us'),
    ('Yuzuru', 'Hanyu', 'jp'),
    ('Ireen', 'Wust', 'nl'),
]
for idx, (fname, lname, code) in enumerate(athletes, 2):
    ws2[f'A{idx}'] = fname
    ws2[f'B{idx}'] = lname
    ws2[f'C{idx}'] = code
wb2.save('athletes_test.xlsx')

# Create results Excel
wb3 = Workbook()
ws3 = wb3.active
ws3['A1'] = 'athleteFirstName'
ws3['B1'] = 'athleteLastName'
ws3['C1'] = 'rank'
ws3['D1'] = 'timeOrPoints'
ws3['E1'] = 'scoreType'
ws3['F1'] = 'medal'
results = [
    ('Katie', 'Ledecky', 1, '3:59.34', 'TIME', 'GOLD'),
    ('Michael', 'Phelps', 2, '4:01.12', 'TIME', 'SILVER'),
    ('Simone', 'Biles', 1, '15.600', 'PTS', 'GOLD'),
    ('Nadia', 'Comaneci', 2, '15.450', 'PTS', 'SILVER'),
    ('Usain', 'Bolt', 1, '9.63', 'TIME', 'GOLD'),
    ('Serena', 'Williams', 1, '1', 'WINS', 'GOLD'),
    ('LeBron', 'James', 2, '2', 'WINS', 'SILVER'),
    ('Cristiano', 'Ronaldo', 1, '750', 'PTS', 'GOLD'),
    ('Lionel', 'Messi', 2, '745', 'PTS', 'SILVER'),
    ('Maria', 'Sharapova', 1, '6', 'WINS', 'GOLD'),
]
for idx, (fname, lname, rank, time_pts, score_type, medal) in enumerate(results, 2):
    ws3[f'A{idx}'] = fname
    ws3[f'B{idx}'] = lname
    ws3[f'C{idx}'] = rank
    ws3[f'D{idx}'] = time_pts
    ws3[f'E{idx}'] = score_type
    ws3[f'F{idx}'] = medal
wb3.save('results_test.xlsx')

print('All Excel files created successfully!')

