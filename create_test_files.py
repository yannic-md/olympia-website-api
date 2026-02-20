#!/usr/bin/env python3
"""
Generate test Excel files for the Excel Importer
"""

import openpyxl
from pathlib import Path
import sys

def create_test_files():
    """Create all test Excel files"""

    # Create test_data directory
    test_dir = Path("test_data")
    test_dir.mkdir(exist_ok=True)

    print("Erstelle Test-Dateien...\n")

    # 1. Countries Sample
    print("1️⃣  Erstelle countries_sample.xlsx...")
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Countries"
    ws.append(["code", "name"])
    ws.append(["USA", "United States"])
    ws.append(["GER", "Germany"])
    ws.append(["FRA", "France"])
    ws.append(["JPN", "Japan"])
    ws.append(["CHN", "China"])
    ws.append(["RUS", "Russia"])
    ws.append(["AUS", "Australia"])

    # Set column widths
    ws.column_dimensions['A'].width = 15
    ws.column_dimensions['B'].width = 25

    wb.save(test_dir / "countries_sample.xlsx")
    print("   ✅ countries_sample.xlsx erstellt")

    # 2. Athletes Sample
    print("2️⃣  Erstelle athletes_sample.xlsx...")
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Athletes"
    ws.append(["firstName", "lastName", "countryCode", "gender"])
    ws.append(["Katie", "Ledecky", "USA", "F"])
    ws.append(["Caeleb", "Dressel", "USA", "M"])
    ws.append(["Michael", "Phelps", "USA", "M"])
    ws.append(["Max", "Mustermann", "GER", "M"])
    ws.append(["Claire", "Dupont", "FRA", "F"])
    ws.append(["Yuki", "Tanaka", "JPN", "F"])
    ws.append(["Sergey", "Petrov", "RUS", "M"])
    ws.append(["Cathy", "Freeman", "AUS", "F"])

    ws.column_dimensions['A'].width = 15
    ws.column_dimensions['B'].width = 15
    ws.column_dimensions['C'].width = 15
    ws.column_dimensions['D'].width = 10

    wb.save(test_dir / "athletes_sample.xlsx")
    print("   ✅ athletes_sample.xlsx erstellt")

    # 3. Results Sample
    print("3️⃣  Erstelle results_sample.xlsx...")
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Results"
    ws.append(["athleteFirstName", "athleteLastName", "rank", "timeOrPoints", "medal"])
    ws.append(["Katie", "Ledecky", 1, "3:59.34", "GOLD"])
    ws.append(["Caeleb", "Dressel", 2, "4:01.12", "SILVER"])
    ws.append(["Michael", "Phelps", 3, "4:02.00", "BRONZE"])
    ws.append(["Max", "Mustermann", 1, "9.85", "GOLD"])
    ws.append(["Claire", "Dupont", 2, "10.15", "SILVER"])
    ws.append(["Yuki", "Tanaka", 3, "10.45", "BRONZE"])
    ws.append(["Sergey", "Petrov", 4, "10.80", ""])
    ws.append(["Cathy", "Freeman", 1, "48.50", "GOLD"])

    ws.column_dimensions['A'].width = 18
    ws.column_dimensions['B'].width = 18
    ws.column_dimensions['C'].width = 10
    ws.column_dimensions['D'].width = 15
    ws.column_dimensions['E'].width = 10

    wb.save(test_dir / "results_sample.xlsx")
    print("   ✅ results_sample.xlsx erstellt")

    # 4. Countries with Errors
    print("4️⃣  Erstelle countries_with_errors.xlsx...")
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Countries"
    ws.append(["code", "name"])
    ws.append(["CAN", "Canada"])  # OK
    ws.append(["USA", "United States"])  # Duplicate
    ws.append(["", "Mexico"])  # Missing code
    ws.append(["BRA", "Brazil"])  # OK
    ws.append(["MEX", ""])  # Missing name
    ws.append(["IND", "India"])  # OK

    ws.column_dimensions['A'].width = 15
    ws.column_dimensions['B'].width = 25

    wb.save(test_dir / "countries_with_errors.xlsx")
    print("   ✅ countries_with_errors.xlsx erstellt")

    # 5. Athletes with Errors
    print("5️⃣  Erstelle athletes_with_errors.xlsx...")
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Athletes"
    ws.append(["firstName", "lastName", "countryCode", "gender"])
    ws.append(["John", "Doe", "USA", "M"])  # OK
    ws.append(["Katie", "Ledecky", "USA", "F"])  # Duplicate
    ws.append(["", "Smith", "GER", "M"])  # Missing firstName
    ws.append(["Jane", "", "FRA", "F"])  # Missing lastName
    ws.append(["Bob", "Johnson", "XXX", "M"])  # Country not found
    ws.append(["Alice", "Williams", "USA", "X"])  # OK (invalid gender accepted)

    ws.column_dimensions['A'].width = 15
    ws.column_dimensions['B'].width = 15
    ws.column_dimensions['C'].width = 15
    ws.column_dimensions['D'].width = 10

    wb.save(test_dir / "athletes_with_errors.xlsx")
    print("   ✅ athletes_with_errors.xlsx erstellt")

    # 6. Results with Errors
    print("6️⃣  Erstelle results_with_errors.xlsx...")
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Results"
    ws.append(["athleteFirstName", "athleteLastName", "rank", "timeOrPoints", "medal"])
    ws.append(["Katie", "Ledecky", 1, "3:59.34", "GOLD"])  # OK
    ws.append(["Unknown", "Athlete", 1, "4:00.00", "GOLD"])  # Athlete not found
    ws.append(["Caeleb", "Dressel", 0, "4:01.12", "SILVER"])  # Invalid rank (0)
    ws.append(["Michael", "Phelps", -1, "4:02.00", ""])  # Invalid rank (negative)
    ws.append(["Max", "Mustermann", 1, "9.85", "PLATINUM"])  # Invalid medal
    ws.append(["Claire", "Dupont", 2, "10.15", "SILVER"])  # OK

    ws.column_dimensions['A'].width = 18
    ws.column_dimensions['B'].width = 18
    ws.column_dimensions['C'].width = 10
    ws.column_dimensions['D'].width = 15
    ws.column_dimensions['E'].width = 10

    wb.save(test_dir / "results_with_errors.xlsx")
    print("   ✅ results_with_errors.xlsx erstellt")

    print("\n" + "="*50)
    print("✨ Alle Test-Dateien wurden erfolgreich erstellt!")
    print("="*50)
    print("\nErstellte Dateien:")
    print("  📁 test_data/")
    print("    ├─ countries_sample.xlsx")
    print("    ├─ athletes_sample.xlsx")
    print("    ├─ results_sample.xlsx")
    print("    ├─ countries_with_errors.xlsx")
    print("    ├─ athletes_with_errors.xlsx")
    print("    └─ results_with_errors.xlsx")
    print("\n✅ Du kannst jetzt die Test-Scripts ausführen!")
    print("\nPowerShell: .\\test_importer.ps1")
    print("Bash:       bash test_importer.sh")

if __name__ == "__main__":
    try:
        create_test_files()
    except Exception as e:
        print(f"\n❌ Fehler beim Erstellen der Test-Dateien:")
        print(f"   {str(e)}")
        sys.exit(1)

